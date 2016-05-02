#!/usr/bin/env python

from __future__ import print_function

import os, sys
import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Style as xlsstyle, Font as xlsfont, Color as xlscolor, Alignment as xlsalign
from openpyxl.comments import Comment as xlscomment

from copy import deepcopy

from miRNexpander.NetworkTools.AliasResolver import AliasResolver


class InteractionLister( AliasResolver ):
    """list out interactions in a specified format"""


    def __init__( self, db_handler ):
        """initialize instance variables"""

        self.Name = None
        self.wb = None

        self._db = db_handler
        self._moltypes = self._db.getConfItem( "moltypes" )

        self._output_handler = None
        self._species_restriction = None
        self._db_restriction = None
        self._reg_type = self._db._reg_type

        self._tdate = re.compile( "[0-9]{4}-[0-9]{2}-[0-9]{2}" )  # date recognition


############################################################
#### PUBLIC                                             ####
############################################################

    def createExcel( self, intseeds, list_dates, tdir, genefilter = None, merge_cells = True ):
        """query the database to build interaction lists"""

        if not intseeds or type( intseeds ) != dict:
            self._extalert( "Parameter 1 is empty or not a dictionary." )
            return False

        if type( genefilter ) in ( tuple, list, dict, set ):
            genefilter = {  g.lower( ) for g in genefilter  }
        elif type( genefilter ) == str:
            genefilter = set( [ genefilter.lower( ) ] )
        else:
            genefilter = None

        # we need to store whether a valid filter was specified for later because genefilter will be reused
        filtered = genefilter != None
        if filtered:
            self._spill( "Filtering with {:d} gene(s)." . format( len( genefilter ) ) )

        # TODO:
        # - database update from new sources
        # - biomolecule hierarchy implementation (NOTE: use URL-like strings)

        urls = {
                "ref" : "https://www.ncbi.nlm.nih.gov/gene/{ncbigene}",
                "out" : {
                    "mirna" : "http://www.mirbase.org/cgi-bin/mirna_entry.pl?acc={mirbase}",
                    "pri-mirna" : "http://www.mirbase.org/cgi-bin/mirna_entry.pl?acc={mirbase}",
                    #" protein" : "http://www.uniprot.org/uniprot/{}",
                    #".protein" : "http://www.uniprot.org/uniprot/?query={}+AND+organism%3Ahuman+AND+reviewed%3Ayes&sort=score",
                    "protein" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "transcriptionfactor" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "targetgene" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "gene" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "transcript" : "",

                    "mrna" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "kinase" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    "phosphoprotein" : "http://www.uniprot.org/uniprot/?query={symbol_uniprot}+AND+organism%3A{species}+AND+reviewed%3Ayes&sort=score",
                    },
                "source" : {
                    "miRTarBase.protein" : "http://mirtarbase.mbc.nctu.edu.tw/php/detail.php?mirtid={}",
                    "miRTarBase.mirna" : "http://mirtarbase.mbc.nctu.edu.tw/php/detail.php?mirtid={}",
                    #"miRTarBase.mirna" : "http://mirtarbase.mbc.nctu.edu.tw/php/search.php?org=hsa&kw={}&opt=mirna_id",
                    #"miRTarBase.protein" : "http://mirtarbase.mbc.nctu.edu.tw/php/search.php?org=hsa&kw={}&opt=target",
                    "TarBase.mirna" : "http://diana.imis.athena-innovation.gr/DianaTools/index.php?r=tarbase/index&mirnas={}",
                    "TarBase.pri-mirna" : "http://diana.imis.athena-innovation.gr/DianaTools/index.php?r=tarbase/index&mirnas={}",
                    "TarBase.protein" : "http://diana.imis.athena-innovation.gr/DianaTools/index.php?r=tarbase/index&genes={}",
                    "miRecords" : "http://mirecords.biolead.org/interactions.php?species=Homo+sapiens&mirna_acc={}&targetgene_type=symbol&targetgene_info={}&v=yes",
                    }
            }

        head_dict = { }
        headers = [ "No", "Symbol", "Data Source", "Pubmed ID", 1, 2, "validated?", "Comment" ]
        for i in xrange( len( headers ) ):
            head_dict[ headers[ i ] ] = {
                                        "icol" : i + 1,
                                        "halign" : "center",
                                        }
        head_dict[ "No" ][ "halign" ] = "right"

        # attrib: "linkouts" combines column headers with the respective data source link type
        attrib = {
                "intranslation.src" : {
                                        "tag" : "has as mRNA targets",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "uniprot", "hgnc.symbol" ) ),
                                        },
                "intranslation.tgt" : {
                                        "tag" : "has as miRNA regulators",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "miRBase", "mirbase" ) ),
                                        },
                "transcription.src" : {
                                        "tag" : "has as gene targets",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "uniprot", "hgnc.symbol" ) ),
                                        },
                "transcription.tgt" : {
                                        "tag" : "has as TF regulators",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "uniprot", "hgnc.symbol" ) ),
                                        },
                "phosphorylation.src" : {
                                        "tag" : "has as phosphoprotein targets",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "uniprot", "hgnc.symbol" ) ),
                                        },
                "phosphorylation.tgt" : {
                                        "tag" : "has as kinase regulators",
                                        "linkouts" : ( ( "NCBI Gene", "ncbigene"), ( "uniprot", "hgnc.symbol" ) ),
                                        },
        }
        defaultstyle = xlsstyle(
                alignment = xlsalign(
                    #vertical = xlsalign.VERTICAL_CENTER,
                    #horizontal = xlsalign.HORIZONTAL_JUSTIFY,
                    vertical = "center",
                    horizontal = "center",
                    )
                )
        hyperstyle = defaultstyle.copy(
                font = xlsfont(
                    color = xlscolor( 'FF0000FF' )
                    )
                )
        row1style = defaultstyle.copy(
                #alignment = xlsalign( horizontal = xlsalign.HORIZONTAL_LEFT )
                alignment = xlsalign( horizontal = "left" )
                )
        col1style = defaultstyle.copy(
                #alignment = xlsalign( horizontal = xlsalign.HORIZONTAL_RIGHT )
                alignment = xlsalign( horizontal = "right" )
                )
        headerstyle = defaultstyle.copy(
                font = xlsfont( bold = True )
                )

        workbooks = { }
        retrieved = { }
        included = { }
        total_edges = { }
        included_edges = { }
        hypercells = [ ]  # keep track of cells with hyperlinks - makes formatting all of them easier

        # loop over the given interactions to retrieve interactors for the given molecules
        for sinter, molecules in intseeds.iteritems( ):
            inter, role = sinter.split( "." )
            try:
                typeset = self._db.getConfItem( "interdefs" )[ inter ][ role ]
            except KeyError:
                self._alert( "Key {!r} is not part of an interaction definition (here {!r}) - use 'src', 'tgt', or 'two-way'." . format( role, inter ) )
                continue

            # preparte headers for spreadsheet (linkouts depend on the type of interaction)
            headers[ 4:6 ] = zip( *attrib[ sinter ][ "linkouts" ] )[ 0 ]
            link_headers = headers[ 4:6 ]
            head_dict.update( zip( link_headers, [ head_dict[ i ] for i in ( 1, 2 ) ] ) )

            # the interaction role affects how results are transformed
            if role == "src":
                keypos = 0,
                valpos = 1,
            elif role == "tgt":
                keypos = 1,
                valpos = 0,
            else:
                # force value for role
                role = "two-way"
                keypos = 0, 1
                valpos = 1, 0

            # find the unique IDs for the given molecules
            refs = self.unalias( molecules, restrict = { "alias_types" : [ "hgnc.symbol" ] }, silent = True )  # NOTE: restriction arbitrary ATM
            symbol_recovery = set( refs )

            # retrieve interactors from the database
            res = self._db.query_for_interactions( refs, {  "ints" : [ inter ], "role" : [ role ]  } )
            if len( res ) == 0:
                self._spill( "No interactions found for input {}." . format( sinter ) )
                continue

            # assign the source and target types to spreadsheet and entry types (ws_type and e_type, respectively)
            ws_type = res[ 0 ][ keypos[ 0 ] + 2 ]
            e_type = res[ 0 ][ valpos[ 0 ] + 2 ]

            # initialize dict that stores interactions, with keys being the input molecules
            entries = dict( [ ( m, { } ) for m in refs ] )

            # transform retrieved interaction results to entries
            for r in res:
                for k, v in zip( keypos, valpos ):
                    try:
                        entries[ r[ k ] ][ r[ v ] ].append( r[ 4 ] )
                    except KeyError:
                        if r[ k ] not in entries:
                            self._alert( "Unexpected result:", r )
                        entries[ r[ k ] ][ r[ v ] ] = [ r[ 4 ] ]

                    try:
                        symbol_recovery |= set( [ r[ 0 ], r[ 1 ] ] )
                    except KeyError:
                        symbol_recovery = set( [ r[ 0 ], r[ 1 ] ] )

            # retrieve annotations
            annot = self.annotate( {  "ids" : symbol_recovery  }, silent = False )
            if not filtered:
                genefilter = {  annot[ i ][ "symbol" ].lower( ) for i in annot  }

            for k, v in self.aliases( symbol_recovery, ( "synonym", ) + zip( *attrib[ sinter ][ "linkouts" ] )[ 1 ], silent = True ).items( ):
                annot[ k ].update( v )

            # self.annotate yields lists that need to be unlisted before they can be used in the format( ) command for linking
            links = deepcopy( annot )
            for k in links:
                links[ k ][ "symbol_uniprot" ] = "+OR+" . join( links[ k ][ "symbol" ].split( ":" ) )
                # consolidate all Entrez IDs into comma-separated list (NCBI website can resolve this)
                links[ k ][ "ncbigene" ] = "," . join( links[ k ][ "ncbigene" ] )
                # consolidate synonyms into readable enumeration for comments
                if "synonym" in links[ k ]:
                    links[ k ][ "synonym" ] = ", " . join( sorted( links[ k ][ "synonym" ] ) )
                # take first list element from everything else
                for ann_type in set( zip( *attrib[ sinter ][ "linkouts" ] )[ 1 ] ) - set( [ "ncbigene", "symbol" ] ):
                    try:
                        links[ k ][ ann_type ] = links[ k ][ ann_type ][ 0 ]
                    except KeyError:
                        continue

            # intermediate function
            def symb( ref ):
                """return the symobl of the given ref, or a pre-defined value if ref is not found"""
                try:
                    return annot[ ref ][ "symbol" ]
                except KeyError:
                    return ref

            # create workbook
            wbname = sinter
            wb = Workbook( )
            wb.remove_sheet( wb.get_active_sheet( ) )
            wb.shared_styles[ 0 ] = defaultstyle
            workbooks[ wbname ] = wb
            total_edges[ wbname ] = 0
            included_edges[ wbname ] = 0

            # prepare containers
            ws_counter = { }
            retrieved[ wbname ] = set( )
            included[ wbname ] = set( )

            ws_intless = wb.create_sheet( title = "<orphaned hooks>" )
            ws_intless.append( [ "Members of the search list without valid interactions:" ] )
            ws_intless.append( [ "Gene symbol", "Description", "Filter Count", "Filtered interactors" ] )
            ws_filtered = wb.create_sheet( title = "<filtered interactors>" )

            for listed_id, interactors in sorted( entries.iteritems( ), key = lambda tup2: symb( tup2[ 0 ] ) ):

                int_symbols = dict( [  ( symb( i ), i ) for i in interactors if i in annot  ] )
                included_ints = dict( [  ( k, v ) for k, v in int_symbols.items( ) if set( k.lower( ).split( ":" ) ) & genefilter  ] )
                retrieved[ wbname ] |= set( int_symbols.values( ) )
                included[ wbname ] |= set( included_ints.values( ) )
                total_edges[ wbname ] += len( int_symbols )
                included_edges[ wbname ] += len( included_ints )
                if not included_ints:
                    # if there are no valid interactions, we add this item to the "interactionless" sheet and continue with the next
                    ws_intless.append( [  symb( listed_id ), annot[ listed_id ][ "description" ], len( int_symbols ), ", " . join( int_symbols.keys( ) )  ] )
                    c = ws_intless.cell( row = ws_intless.get_highest_row( ), column = 1 )
                    c.hyperlink = urls[ "ref" ].format( **links[ listed_id ] )
                    if "synonym" in links[ listed_id ]:
                        c.comment = xlscomment( links[ listed_id ][ "synonym" ], "[automatic]" )
                    hypercells.append( c )
                    continue

                # create and initialize the sheet
                input_id = symb( listed_id ) + " (" + refs[ listed_id ] + ")"
                ws = wb.create_sheet( title = input_id )
                ws.append( [ input_id, attrib[ sinter ][ "tag" ] ] )  # row 1
                c = ws.cell( "A1" )
                if "synonym" in links[ listed_id ]:
                    c.comment = xlscomment( links[ listed_id ][ "synonym" ], "[automatic]" )
                try:
                    c.hyperlink = urls[ "ref" ].format( **links[ listed_id ] )
                except KeyError:
                    pass
                hypercells.append( c )
                ws.append( headers )  # row 2
                for row in ws.range( "A2:" + chr( 65 + len( headers ) ) + "2" ):
                    for cell in row:
                        cell.style = headerstyle
                ws_counter[ input_id ] = 0

                # write interactors
                for symbol, id2 in sorted( included_ints.items( ) ):
                    ws_counter[ input_id ] += 1
                    instances = interactors[ id2 ]

                    try:
                        outlink1 = links[ id2 ][ attrib[ sinter ][ "linkouts" ][ 0 ][ 1 ] ]
                    except KeyError:
                        outlink1 = None
                    try:
                        outlink2 = links[ id2 ][ attrib[ sinter ][ "linkouts" ][ 1 ][ 1 ] ]
                    except KeyError:
                        outlink2 = None

                    last_db = None
                    start_rows = [ ]
                    for i in instances:
                        try:  # assume comma-separated string
                            pmids = i[ "PMIDs" ].split( "," )
                        except AttributeError:  # happens when this is a single entry
                            pmids = [ str( i[ "PMIDs" ] ) ]
                        intlink = i[ "accession" ] if "accession" in i else symbol

                        ws.append( [
                            ws_counter[ input_id ],  # No
                            symbol,  # Symbol
                            i[ "database" ] + " r" + i[ "release" ],  # Data Source
                            pmids[ 0 ],  # PMIDs
                            outlink1,
                            outlink2,
                            ]
                        )

                        if i[ "database" ] != last_db:
                            start_rows.append( ws.get_highest_row( ) )
                            last_db = i[ "database" ]

                        current_row = ws.get_highest_row( )

                        # link pubmed identifier
                        c = ws.cell( row = current_row, column = head_dict[ "Pubmed ID" ][ "icol" ] )
                        c.hyperlink = "https://www.ncbi.nlm.nih.gov/pubmed/" + pmids[ 0 ]
                        hypercells.append( c )

                        # link to data source
                        c = ws.cell( row = current_row, column = head_dict[ "Symbol" ][ "icol" ] )
                        try:
                            c.comment = xlscomment( links[ id2 ][ "synonym" ], "[automatic]" )
                        except KeyError:
                            pass
                        try:
                            c.hyperlink = urls[ "source" ][ i[ "database" ] + "." + e_type ].format( intlink )
                            hypercells.append( c )
                        except KeyError:
                            pass

                        # link outlinks (Uniprot, Ensembl, miRBase, ... identifiers)
                        c = ws.cell( row = current_row, column = head_dict[ link_headers[ 0 ] ][ "icol" ] )
                        if c.value:
                            c.hyperlink = urls[ "ref" ].format( **links[ id2 ] )
                            hypercells.append( c )

                        if outlink2:
                            c = ws.cell( row = current_row, column = head_dict[ link_headers[ 1 ] ][ "icol" ] )
                            c.hyperlink = urls[ "out" ][ e_type ].format( **links[ id2 ] )
                            hypercells.append( c )

                        # add and link additional pubmed identifers
                        for p in pmids[ 1: ]:
                            ws.append( [ None, None, None, p ] )
                            current_row += 1
                            c = ws.cell( row = current_row, column = head_dict[ "Pubmed ID" ][ "icol" ] )
                            c.hyperlink = "https://www.ncbi.nlm.nih.gov/pubmed/" + p
                            hypercells.append( c )

                    # NOTE: merging prevents sorting in Excel from working
                    start_rows[ 0 ] = str( start_rows[ 0 ] )
                    ws.merge_cells( "A" + start_rows[ 0 ] + ":A" + str( current_row ) )

                    # merge other cells (if asked to do so)
                    if merge_cells:
                        # one merger across the whole interactor
                        for column in "E", "F":
                            ws.merge_cells( column + start_rows[ 0 ] + ":" + column + str( current_row ) )

                        # separate merges for databases
                        for column in "B", "C":
                            i = 0
                            while i < len( start_rows ) - 1:
                                ws.merge_cells( column + str( start_rows[ i ] ) + ":" + column + str( start_rows[ i + 1 ] - 1 ) )
                                i += 1
                            ws.merge_cells( column + str( start_rows[ i ] ) + ":" + column + str( current_row ) )

                if ws.cell( "C1" ).value == None:
                    ws.cell( "C1" ).value = str( ws_counter[ input_id ] ) + ( " interactor in total" if ws_counter[ input_id ] == 1 else " interactors in total" )
                    first_row = "A1", "B1", "C1"
                else:
                    ws.cell( "E1" ).value = str( ws_counter[ input_id ] ) + ( " interactor in total" if ws_counter[ input_id ] == 1 else " interactors in total" )
                    first_row = "A1", "B1", "C1", "D1", "E1"

                # left-justify the very first row
                for c in first_row:
                    ws.cell( c ).style = row1style

            # output rejected interarctors into separate sheet
            if filtered:
                interactors = {  ( annot[ gene_id ][ "symbol" ], gene_id ) for gene_id in ( retrieved[ wbname ] - included[ wbname ] ) & set( annot )  }
                ws_filtered.append( [ "Identifiable interactors filtered as irrelevant:", str( len( interactors ) ) ] )
                ws_filtered.append( [ "Gene symbol", "Description" ] )
                current_row = 2
                for gene in sorted( interactors ):
                    ws_filtered.append( [ gene[ 0 ], annot[ gene[ 1 ] ][ "description" ] ] )
                    current_row += 1
                    c = ws_filtered.cell( row = current_row, column = 1 )
                    c.hyperlink = urls[ "ref" ].format( **links[ gene[ 1 ] ] )
                    if "synonym" in links[ gene[ 1 ] ]:
                        c.comment = xlscomment( links[ gene[ 1 ] ][ "synonym" ], "[automatic]" )
                    hypercells.append( c )
            else:
                ws_filtered.append( [ "No filtering took place during the creation of this file." ] )

            ws_intless.cell( "B1" ).value = ws_intless.get_highest_row( ) - 2

        # apply defined style to hyperlinked cells
        for c in hypercells:
            c.style = hyperstyle

        # finishing touches and writing to disk
        filenames = self._db.getConfItem( "inter_fn" )
        d = {
                "filtered" : "filtered_" if filtered else "",
                "rdate" : date.today( ).strftime( "%Y-%m-%d" ),
                "species" : "+" . join( map( str, self._db.getAllowedSpecies( ) ) ),
        }
        for wbname, wb in sorted( workbooks.items( ) ):
            d.update( {  "ldate" : "+" . join( sorted( list_dates[ wbname ] ) )  } )
            for wsname in wb.get_sheet_names( ):
                ws = wb.get_sheet_by_name( wsname )
                ws.freeze_panes = "B3"
                # set column width
                for col in ws.columns:
                    column_width = max( [ len( str( cell.value ) ) for cell in col ] )
                    ws.column_dimensions[ cell.column ].width = 1.3 * column_width

            fn = filenames[ wbname ] . format( **d ) + ".xlsx"
            wb.save( os.path.join( tdir, fn ) )
            self._spill( ( "Filtered {:d} ({:d}) of {:d} ({:d}) interactor(s) to " + os.path.join( tdir, fn ) ) .
                            format( len( included[ wbname ] ), included_edges[ wbname ], len( retrieved[ wbname ] ), total_edges[ wbname ] ) )

        return True


    def readIntseeds( self, filenames ):
        """reads and transforms input from a file into an internal intseeds data structure to be used in .createlLists"""

        intseeds = { }
        list_dates = { }
        for f in filenames:
            list_date = self._tdate.search( os.path.basename( f ) )
            if not list_date:
                self._alert( "Unable to extract list creation date (YYYY-MM-DD) from filename {!r}." . format( f ) )
                while not list_date:
                    helper = raw_input( "Please enter the list creation date in the format specified above (empty input skips this file): " )
                    if helper == "":  # skip the file
                        break
                    else:
                        list_date = self._tdate.search( helper )
                if not helper:  # skip the file
                    continue
            list_date = list_date.group( )

            try:
                stream = open( f )
            except IOError:
                self._alert( "Unable to read {!r}, skipping file." . format( f ) )
                continue

            section = None
            sectionless = False
            for line in stream:
                if line.strip( ) == "" or line.strip( )[ 0 ] == "#":  # ignore emtpy lines and comments
                    continue
                elif line[ 0 ] != "\t":  # lines not beginning with tab start a new section
                    section = line.strip( ).split( "." )
                    if section[ 0 ] not in self._db.getConfItem( "interdefs" ):
                        self._alert( "Section {!r} is not recognized as an interaction definition, aborting." . format( line.strip( ) ) )
                        return

                    if len( section ) == 1:
                        section.append( "two-way" )
                    elif section[ 1 ] not in ( "src", "tgt" ):
                        self._alert( "Section {!r} has an invalid subsection: {!r}, aborting." . format( line.strip( ), section[ 1 ] ) )
                        return

                    section = "." . join( section )
                    #if not section in intseeds:
                    #    intseeds[ section ] = dict( )
                    #    list_dates[ section ] = dict( )
                else:  # all other lines are expected to contain members of the most recent section
                    if not section:
                        sectionless = True
                        continue

                    try:
                        #intseeds[ section[ 0 ] ][ section[ 1 ] ].add( line.strip( ) )
                        #list_dates[ section[ 0 ] ][ section[ 1 ] ].add( list_date )
                        intseeds[ section ].add( line.strip( ) )
                        list_dates[ section ].add( list_date )
                    except KeyError:
                        #intseeds[ section[ 0 ] ][ section[ 1 ] ] = set( [ line.strip( ) ] )
                        #list_dates[ section[ 0 ] ][ section[ 1 ] ] = set( [ list_date ] )
                        intseeds[ section ] = set( [ line.strip( ) ] )
                        list_dates[ section ] = set( [ list_date ] )

            if sectionless:
                self._alert( "There is a missing section name at the beginning of file {!r}, entries were discarded." . format( f ) )

        if intseeds == { }:
            intseeds = None
        return intseeds, list_dates



############################################################
#### NOT SO PUBLIC                                      ####
############################################################

